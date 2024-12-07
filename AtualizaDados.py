from openpyxl import load_workbook
from copy import copy

# Função para converter a tupla do openpyxl em uma array
def convert_to_array(data_tuple):
    return [cell.value for cell in data_tuple]

# Função para identificar os índices das colunas que não estão na planilha 1 mas estão na 2
def find_indexes(uniques, arr):
    new_arr = []
    for unique in uniques:
        new_arr = new_arr + [i for i,n in enumerate(arr) if n==unique]
    return new_arr

# Entrada dos dados pelo usuário
spreadsheet1 = input("Nome da planilha que receberá os dados: ")
spreadsheet2 = input("Nome da planilha que contém os dados: ")
colkey = input("Coluna chave de transferência: ")

# Adicionando a extensão do arquivo planilha excel
filename1 = spreadsheet1 + '.xlsx'
filename2 = spreadsheet2 + '.xlsx'

# Carregando as planilhas com os nomes de input
wb1 = load_workbook(filename=filename1)
wb2 = load_workbook(filename=filename2)

# Configurando a pasta de trabalho como a pasta ativa
ws1 = wb1.active
ws2 = wb2.active

# Extraindo os nomes das colunas de dados presentes nas duas planilhas
headers1 = convert_to_array(ws1['1'])
headers2 = convert_to_array(ws2['1'])

# Extraindo as colunas únicas que estão na planilha 2 e não estão na 1
uniques = list(set(headers2) - set(headers1))

# Encontrar quais são os índices dessas colunas únicas
indices = find_indexes(uniques, headers2)

# Array que contém os nomes únicos da coluna KEY
keys = convert_to_array(ws1[colkey])

for col_idx in indices:
    col_num = col_idx + 1
    col_data = ws2[col_num]
    for row in range(1, ws2.max_row + 1):
        key = ws2.cell(row=row,column=1).value
        if key in keys:
            rowidx = keys.index(key) + 1
            value = ws2.cell(row=row,column=col_num).value
            ws1.cell(row=rowidx,column=col_num).value = value
            if ws2.cell(row=row,column=col_num).has_style:
                ws1.cell(row=rowidx,column=col_num).font = copy(ws2.cell(row=row,column=col_num).font)
                ws1.cell(row=rowidx,column=col_num).border = copy(ws2.cell(row=row,column=col_num).border)
                ws1.cell(row=rowidx,column=col_num).fill = copy(ws2.cell(row=row,column=col_num).fill)
                ws1.cell(row=rowidx,column=col_num).number_format = copy(ws2.cell(row=row,column=col_num).number_format)
                ws1.cell(row=rowidx,column=col_num).protection = copy(ws2.cell(row=row,column=col_num).protection)
                ws1.cell(row=rowidx,column=col_num).alignment = copy(ws2.cell(row=row,column=col_num).alignment)

# Salvando os arquivos editados
wb1.save(spreadsheet1 + "_ATT" + ".xlsx")
# wb2.save(spreadsheet2 + "_ATT" + ".xlsx")
